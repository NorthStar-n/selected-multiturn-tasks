"""Build the numeric analysis workbook using only the governed register workbook."""

from __future__ import annotations

import hashlib
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[3]
OUTPUT = ROOT / "output"
REGISTER = OUTPUT / "weekly_planning_review_register.xlsx"
ANALYSIS = OUTPUT / "weekly_planning_numeric_analysis.xlsx"


def sheet_rows(workbook, name: str) -> list[dict[str, object]]:
    sheet = workbook[name]
    values = list(sheet.iter_rows(values_only=True))
    return [dict(zip(values[0], row)) for row in values[1:]]


def add_sheet(workbook: Workbook, title: str, fields: list[str], rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    fill = PatternFill("solid", fgColor="7030A0")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    table = Table(displayName=title.replace("_", ""), ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2, 55
        )


def to_date(value: object, day_first: bool = False) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value)
    return datetime.strptime(text, "%d/%m/%Y" if day_first else "%Y-%m-%d").date()


def add_workdays(start: date, count: int, non_working: set[date]) -> date:
    current = start
    added = 0
    while added < count:
        current += timedelta(days=1)
        if current.weekday() < 5 and current not in non_working:
            added += 1
    return current


def trade_analysis(register) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    source = sheet_rows(register, "Source_Records")
    ppi = {int(r["year"]): float(r["annual_ppi_average"]) for r in sheet_rows(register, "PPI_Annual")}
    macro = {int(r["year"]): r for r in sheet_rows(register, "Macro_Context")}
    values: dict[int, tuple[float, str, str]] = {}
    for row in source:
        if row["flow_code"] != "X":
            continue
        if row["inclusion_status"] == "included_headline":
            values[int(row["ref_year"])] = (float(row["primary_value"]), "official_catalog", str(row["record_id"]))
        elif row["inclusion_status"] == "included_bridge":
            values[int(row["ref_year"])] = (float(row["primary_value"]), "modeled_export_bridge", str(row["record_id"]))

    components = ["exports_gdp_pct", "manufacturing_value_added", "chemical_employment_thousand"]
    base = {component: float(macro[2022][component]) for component in components}
    macro_index = {
        year: sum(float(row[c]) / base[c] * 100 for c in components) / len(components)
        for year, row in macro.items()
    }
    years = list(range(2015, 2023))
    real = {year: values[year][0] / ppi[year] * 100 for year in years}
    x = np.array([[1.0, year - 2015, macro_index[year]] for year in years])
    y = np.array([real[year] for year in years])
    coefficients, _, _, _ = np.linalg.lstsq(x, y, rcond=None)
    fitted = x @ coefficients
    ss_res = float(np.sum((y - fitted) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r_squared = 1 - ss_res / ss_tot

    for year in (2023, 2024):
        predicted_real = float(np.array([1.0, year - 2015, macro_index[year]]) @ coefficients)
        predicted_nominal = predicted_real * ppi[year] / 100
        values[year] = (predicted_nominal, "regression_forecast", f"FORECAST-GBR-HS29-{year}")
        real[year] = predicted_real

    baseline = values[2019][0]
    trade_rows = []
    for year in range(2015, 2025):
        nominal, classification, record_id = values[year]
        ratio = nominal / baseline
        trade_rows.append({
            "year": year,
            "record_id": record_id,
            "source_classification": classification,
            "nominal_export_usd": nominal,
            "annual_ppi_average": ppi[year],
            "real_volume_index_value": real[year],
            "macro_index_2022_eq_100": macro_index[year],
            "recovery_ratio_vs_2019": ratio,
            "meets_92pct_recovery_benchmark": ratio >= 0.92,
            "governing_rule": "REV-MR-23-03; MODEL-BRIDGE-23-03",
        })
    model_rows = [
        {"metric": "intercept", "value": float(coefficients[0]), "definition": "OLS coefficient"},
        {"metric": "year_offset", "value": float(coefficients[1]), "definition": "OLS coefficient; year minus 2015"},
        {"metric": "macro_index", "value": float(coefficients[2]), "definition": "OLS coefficient; component-normalized macro index"},
        {"metric": "r_squared", "value": r_squared, "definition": "Training fit for 2015-2022 bridged real volume"},
        {"metric": "recovery_benchmark", "value": 0.92, "definition": "REV-MR-23-03"},
        {"metric": "pre_shock_baseline_year", "value": 2019, "definition": "Last pre-2020 bridge year"},
    ]
    return trade_rows, model_rows


def procurement_analysis(register) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    orders = sheet_rows(register, "Procurement_Orders")
    receipts = sheet_rows(register, "Procurement_Receipts")
    materials = {r["erp_code"]: r for r in sheet_rows(register, "Material_Catalog")}
    pathways = sheet_rows(register, "Supply_Pathways")
    rules = sheet_rows(register, "Operational_Rules")
    entities = {r["erp_code"]: r for r in sheet_rows(register, "Entity_Resolution")}

    threshold = float(next(r["numeric_value"] for r in rules if r["parameter"] == "review_threshold"))
    reserve = float(next(r["numeric_value"] for r in rules if r["parameter"] == "late_expedite_reserve"))
    non_working = {
        datetime.strptime(str(int(r["numeric_value"])), "%Y%m%d").date()
        for r in rules if r["parameter"] == "non_working_date"
    }
    received: dict[str, float] = defaultdict(float)
    for receipt in receipts:
        received[str(receipt["po_number"])] += float(receipt["qty_kg"])
    dependency: dict[str, float] = defaultdict(lambda: 1.0)
    tier3: dict[str, str] = {}
    for path in pathways:
        erp = str(path["erp_code"])
        dependency[erp] *= float(path["dependency_ratio"])
        if path["tier_step"] == "tier3_to_tier2":
            tier3[erp] = str(path["child_supplier"])

    po_rows: list[dict[str, object]] = []
    for order in orders:
        po = str(order["po_number"])
        erp = str(order["erp_code"])
        ordered = float(order["ordered_qty_kg"])
        received_qty = received[po]
        open_qty = max(ordered - received_qty, 0.0)
        dispatch = to_date(order["dispatch_date"])
        stated_target = to_date(order["dock_target_date"])
        governed_target = add_workdays(dispatch, int(order["lead_time_working_days"]), non_working)
        late = governed_target > stated_target
        unit_cost = float(materials[erp]["unit_cost_gbp_per_kg"])
        base = open_qty * unit_cost
        dep_adjusted = base * dependency[erp]
        reserve_gbp = dep_adjusted * reserve if late and open_qty > 0 else 0.0
        po_rows.append({
            "po_number": po,
            "canonical_entity_id": entities[erp]["canonical_entity_id"],
            "erp_code": erp,
            "supplier_portal_id": entities[erp]["supplier_portal_id"],
            "bom_reference": entities[erp]["bom_reference"],
            "hs_tariff_code": entities[erp]["hs_tariff_code"],
            "trade_partner_market_code": entities[erp]["trade_partner_market_code"],
            "ordered_qty_kg": ordered,
            "received_qty_kg": received_qty,
            "open_qty_kg": open_qty,
            "unit_cost_gbp_per_kg": unit_cost,
            "dependency_ratio": dependency[erp],
            "base_open_value_gbp": base,
            "dependency_adjusted_exposure_gbp": dep_adjusted,
            "stated_dock_target": stated_target,
            "governed_dock_target": governed_target,
            "late_after_calendar_rules": late,
            "expedite_reserve_gbp": reserve_gbp,
            "risk_adjusted_exposure_gbp": dep_adjusted + reserve_gbp,
        })

    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in po_rows:
        grouped[str(row["erp_code"])].append(row)
    risk_rows = []
    for erp, items in sorted(grouped.items()):
        exposure = sum(float(row["risk_adjusted_exposure_gbp"]) for row in items)
        entity = entities[erp]
        above = exposure >= threshold
        risk_rows.append({
            "risk_id": f"RISK-{erp}",
            "canonical_entity_id": entity["canonical_entity_id"],
            "erp_code": erp,
            "po_numbers": "; ".join(str(row["po_number"]) for row in items),
            "supplier_portal_id": entity["supplier_portal_id"],
            "bom_reference": entity["bom_reference"],
            "hs_tariff_code": entity["hs_tariff_code"],
            "trade_partner_market_code": entity["trade_partner_market_code"],
            "open_qty_kg": sum(float(row["open_qty_kg"]) for row in items),
            "tier3_source": tier3.get(erp, "none identified"),
            "risk_adjusted_exposure_gbp": exposure,
            "review_threshold_gbp": threshold,
            "threshold_exceeded": above,
            "action_status": "review_required" if above else "below_threshold",
            "exception_document_id": "CR-2219 (expired 2023-01-06)" if erp == "ERP-OC-1042" else "",
            "governing_rule": "REV-MR-23-03 section 3; FIN-RES-22-118; OPS-CAL-2022",
        })
    return po_rows, risk_rows


def build() -> None:
    if not REGISTER.exists():
        raise FileNotFoundError(f"Governed register is required: {REGISTER}")
    register_hash = hashlib.sha256(REGISTER.read_bytes()).hexdigest()
    register = load_workbook(REGISTER, read_only=True, data_only=True)
    required = {
        "Source_Records", "Governance", "Entity_Resolution", "PPI_Annual",
        "Macro_Context", "Procurement_Orders", "Procurement_Receipts",
        "Material_Catalog", "Supply_Pathways", "Operational_Rules",
    }
    missing = required.difference(register.sheetnames)
    if missing:
        raise ValueError(f"Register is missing required sheets: {sorted(missing)}")

    trade, model = trade_analysis(register)
    po_analysis, risk = procurement_analysis(register)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Read_Me"
    summary.append(["Weekly planning numeric analysis"])
    summary.append(["Sole upstream input", REGISTER.name])
    summary.append(["Register SHA-256", register_hash])
    summary.append(["Lineage rule", "All numeric inputs are read from the governed register workbook; raw data files are not read by this builder."])
    summary.append(["Trade observations", len(trade)])
    summary.append(["Procurement POs", len(po_analysis)])
    summary.append(["ERP risks", len(risk)])
    summary["A1"].font = Font(size=14, bold=True)
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 95

    add_sheet(workbook, "Trade_Annual", list(trade[0]), trade)
    add_sheet(workbook, "Trade_Model", list(model[0]), model)
    add_sheet(workbook, "Procurement_PO", list(po_analysis[0]), po_analysis)
    add_sheet(workbook, "Continuity_Risk", list(risk[0]), risk)
    methods = [
        {"calculation": "headline_trade_value", "formula_or_rule": "All-world row where partner_code=0, partner2_code=0, and mot_code=0", "status": "governed", "reference": "REV-MR-22-02 section 2 retained by REV-MR-23-03"},
        {"calculation": "real_volume", "formula_or_rule": "nominal_export_usd / annual_ppi_average * 100", "status": "governed", "reference": "MODEL-BRIDGE-23-03"},
        {"calculation": "macro_index", "formula_or_rule": "Mean of exports/GDP, manufacturing value added, and chemical employment after separately normalizing each component to 2022=100", "status": "implementation interpretation", "reference": "MODEL-BRIDGE-23-03"},
        {"calculation": "forecast", "formula_or_rule": "OLS real_volume ~ intercept + (year-2015) + macro_index; train 2015-2022; convert predicted real volume with annual PPI", "status": "governed implementation", "reference": "REV-MR-23-03; MODEL-BRIDGE-23-03"},
        {"calculation": "recovery_ratio", "formula_or_rule": "nominal export / 2019 bridge export", "status": "explicit analysis assumption", "reference": "2019 selected as last pre-2020 shock year; benchmark=92% per REV-MR-23-03"},
        {"calculation": "received_qty", "formula_or_rule": "Sum receipts by PO including negative quality reversals and ASN-GBR-773 late matcher quantity", "status": "governed source handling", "reference": "OPS-RUN-DOCK-07; ASN-GBR-773"},
        {"calculation": "open_qty", "formula_or_rule": "max(ordered_qty - received_qty, 0)", "status": "implementation", "reference": "weekly planning convention"},
        {"calculation": "dependency_ratio", "formula_or_rule": "Product of dependency ratios along the registered pathway for each ERP code", "status": "explicit analysis assumption", "reference": "Supply_Pathways register sheet"},
        {"calculation": "risk_adjusted_exposure", "formula_or_rule": "open_qty * unit_cost * dependency_ratio * (1 + 12% when governed target is later than stated dock target)", "status": "explicit analysis assumption using governed reserve", "reference": "FIN-RES-22-118; OPS-CAL-2022"},
        {"calculation": "threshold_test", "formula_or_rule": "risk_adjusted_exposure >= 35000 GBP at ERP-code level", "status": "governed", "reference": "REV-MR-23-03 section 3"},
    ]
    add_sheet(workbook, "Calculation_Methods", list(methods[0]), methods)
    governance = sheet_rows(register, "Governance")
    add_sheet(workbook, "Governance_Used", list(governance[0]), governance)
    entities = sheet_rows(register, "Entity_Resolution")
    add_sheet(workbook, "Entity_Resolution", list(entities[0]), entities)
    workbook.save(ANALYSIS)


if __name__ == "__main__":
    build()
